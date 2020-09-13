import PySimpleGUI as sg
from openpyxl import load_workbook
from openpyxl import Workbook
from requests import get
from json import loads
import time
from time import gmtime
from time import strftime
import sys

MAX_TRIES = 10

layout = [  
    [sg.Text('Selecione o arquivo (.xlsx)', font='Roboto 10')],
    [sg.In() ,sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),), key='file')],
    [sg.ProgressBar(100, orientation='h', size=(40, 5), key='progressbar')],
    [sg.InputText(size=(15, 3), font='Roboto 10', key='registrosProcessados', readonly=True), sg.Text('Registros processados', font='Roboto 10')],
    [sg.InputText(size=(15, 3), font='Roboto 10', key='tempoEstimado', readonly=True), sg.Text('Tempo estimado', font='Roboto 10')],
    [sg.Button('Buscar dados', key='enviar', font='Roboto 10')]
]

window = sg.Window('gamalu-crawler', layout, margins=(15, 15)).Finalize()

while True:

    event, values = window.read()
    progressBar = window['progressbar']

    if event == sg.WIN_CLOSED or event == 'Cancel': # if user closes window or clicks cancel
        break

    workbook = load_workbook(filename = values['file'])
    worksheet = workbook.worksheets[0]

    registrosTotais = worksheet.max_row

    window['registrosProcessados'].update('0/' + str(registrosTotais))

    wb = Workbook()
    ws = wb.active

    ws.append([

        "reference_row",
        "zip_code",
        "product",
        "description",
        "distribution_center",
        "name",
        "city",
        "sellers",
        "is_deadline",
        "price",
        "time",
        "zip_code_restriction"

    ])

    wb.save("Dados magalu.xlsx")

    for index, row in enumerate(worksheet.iter_rows()):

        startTime = time.time()

        if(index == 0):
            continue

        cep = str(row[0].value)
        product = str(row[1].value)

        if(cep == "None"):
            break

        productURL = 'https://www.magazineluiza.com.br/produto/calculo-frete/'+ cep +'/'+ product +'/magazineluiza.json'
        productHeaders = {
            'user-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36',
            'referer': 'https://www.magazineluiza.com.br',
        }

        productRequest = []

        trying = 0
        while(trying < MAX_TRIES):

            try:
                productRequest = get(productURL, headers=productHeaders, verify=True)
                productRequest.raise_for_status()
                break

            except:
                print('Nao foi possivel realizar a requisicao, tentando novamente... (' + str(trying) + '/' + str(MAX_TRIES) + ')')

            trying += 1

        print(productRequest)
        print(productRequest.text)

        if(trying == MAX_TRIES):
            continue
        
        productData = loads(productRequest.text)

        for deliveryOption in productData['delivery']:

            name = ""
            city = ""
            sellers = ""

            if deliveryOption['distribution_center'] > 0:

                distributionCenterURL = 'https://lojas.magazineluiza.com.br/filiais/' + str(deliveryOption['distribution_center'])
                distributionCenterRequest = []

                trying = 0
                while(trying < MAX_TRIES):

                    try:
                        distributionCenterRequest = get(distributionCenterURL, verify=True)
                        productRequest.raise_for_status()
                        break

                    except:
                        print('Nao foi possivel realizar a requisicao, tentando novamente... (' + str(trying) + '/' + str(MAX_TRIES) + ')')

                    trying += 1

                if(trying == MAX_TRIES):
                    continue
                
                distributionCenterHTML = distributionCenterRequest.text
                distributionCenterHTML.find("col-7 col-md-8 m-auto")

                firstCut = distributionCenterHTML[(103+distributionCenterHTML.find("col-7 col-md-8 m-auto")):]
                secondCut = firstCut[firstCut.find("data-v-12a79897")+16:]
                thirdCut = secondCut[secondCut.find("data-v-12a79897")+67:]

                name = firstCut.split("<")[0].strip()
                city = secondCut.split("<")[0]
                sellers = thirdCut.split("<")[0]
            
            ws.append([

                index,
                cep,
                product,
                deliveryOption['description'],
                deliveryOption['distribution_center'],
                name,
                city,
                sellers,
                deliveryOption['is_deadline'],
                deliveryOption['price'],
                deliveryOption['time'],
                deliveryOption['zip_code_restriction']

            ])

            wb.save("Dados magalu.xlsx")

            endTime = time.time()
            totalTime = strftime("%H:%M:%S", gmtime(((endTime-startTime) * (registrosTotais-index))))

            window['registrosProcessados'].update(str(index) + '/' + str(registrosTotais))
            window['tempoEstimado'].update(totalTime)

            progressBar.UpdateBar( (index/registrosTotais)*100 )

    wb.close()
    sg.popup('Busca finalizada! Arquivo gerado: Dados magalu.xlsx')